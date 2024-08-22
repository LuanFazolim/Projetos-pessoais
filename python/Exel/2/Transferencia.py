import os
import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side
if os.path.exists("carteira de clientes CAMILA TESTE Res.xlsx"):
    os.remove("carteira de clientes CAMILA TESTE Res.xlsx")
arquivo = openpyxl.load_workbook("carteira de clientes CAMILA TESTE.xlsx")
print(arquivo.sheetnames)
pg1 = arquivo["clientes novos "]
pg2 = arquivo["2024 "]
pg3 = arquivo["indicação clientes gerencia "]

font = Font(name='Calibri', size=11, bold=False, color='000000')
alignment = Alignment(horizontal='left')
bordas = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)



lista = list()
cont = 3
#A====================================
for linha in pg1["A"]:
    cont +=1
    if cont == 4:
        lista.append(linha.value)

        cont = 0
y = 2
x = 4
for fds in pg3.iter_cols(min_col=2):
    for fds2 in pg3.iter_rows(min_row=4):
        if pg3.cell(row=x, column=y).value != None:
            print(pg3.cell(row=x, column=y).value, f"   X = {x}  Y = {y}")
            lista.append(pg3.cell(row=x, column=y).value)

        x += 1
    y += 1
    x = 4

for t in range(len(lista)):
    pg2[f"A{t+9}"] = lista[t]
    pg2[f"A{t+9}"].font = font
    pg2[f"A{t + 9}"].alignment = alignment
    pg2[f"A{t + 9}"].border = bordas
    pg2[f"E{t + 9}"] = "-"
    pg2[f"E{t + 9}"].border = bordas
    pg2[f"E{t + 9}"].font = font
    pg2[f"E{t + 9}"].alignment = alignment

    pg2[f"C{t + 9}"] = "-"
    pg2[f"C{t + 9}"].font = font
    pg2[f"C{t + 9}"].alignment = alignment
    pg2[f"C{t + 9}"].border = bordas

    pg2[f"D{t+9}"] = "-"
    pg2[f"D{t+9}"].font = font
    pg2[f"D{t + 9}"].alignment = alignment
    pg2[f"D{t + 9}"].border = bordas

    pg2[f"B{t + 9}"] = "-"
    pg2[f"B{t + 9}"].font = font
    pg2[f"B{t + 9}"].alignment = alignment
    pg2[f"B{t + 9}"].border = bordas
print(lista)

#B=====================================

cont = 1
lista.clear()
for linha in pg1["A"]:
    cont +=1
    if cont == 4:

        lista.append(linha.value)
        cont = 0
for t in range(len(lista)):
    pg2[f"C{t+9}"] = "cotia"
    pg2[f"C{t + 9}"].font = font
    pg2[f"C{t + 9}"].alignment = alignment
    pg2[f"C{t + 9}"].border = bordas

    pg2[f"B{t+9}"] = lista[t]
    pg2[f"B{t+9}"].font = font
    pg2[f"B{t+9}"].alignment = alignment
    pg2[f"B{t + 9}"].border = bordas


lista.clear()
cont = 0
#D=====================================
for linha in pg1["A"]:
    cont +=1
    if cont == 4:

        lista.append(linha.value)

        cont = 0

for t in range(len(lista)):
    pg2[f"D{t+9}"] = lista[t]
    pg2[f"D{t+9}"].font = font
    pg2[f"D{t + 9}"].alignment = alignment
    pg2[f"D{t + 9}"].border = bordas
lista.clear()

#pag 3=========================================




print()
arquivo.save("carteira de clientes CAMILA TESTE Res.xlsx")

os.startfile("carteira de clientes CAMILA TESTE Res.xlsx")




