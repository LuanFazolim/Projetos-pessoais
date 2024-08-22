
import openpyxl

arquivo = openpyxl.load_workbook("carteira de clientes CAMILA.xlsx")
pagina = arquivo["clientes completa"]
print(arquivo.sheetnames)

titulo = list()
cont = 1
contg = 0

for linha in pagina.iter_rows(min_row=5):
    if cont == 1:
        while cont < 9:
            titulo.append(linha[cont].value)
            cont +=1
        cont = 0
    else:
        contg += 1
        print(f"==== Cliente {contg} ====")
        while cont < 8:

            print(f"{titulo[cont]}:",linha[cont+1].value)
            cont+=1
        print()
        cont = 0




